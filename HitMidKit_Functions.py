import pandas as pd
import numpy as np
import os
import daxQueries as daxQ
import datetime
import time
from openpyxl import load_workbook
import xlwings as xw
import sys
from sys import path
#from openpyxl import load_workbook
#path.append("c:\Mariusz\MyProjects\HitMidKit_Downloader\dll")
#path.append(os.getcwd() + '\\dll')

if getattr(sys, 'frozen', False):
    path.append(os.path.dirname(sys.executable) + '\\dll')
else:
    path.append(os.path.dirname(os.path.abspath(__file__)) + '\\dll')

from pyadomd import Pyadomd

"""
Hit Mid Kit - Inventory Downloader
"""
def dataFrameFromTabular(query):
    """
    # w path.append jest istotna ścieżka
    # muszą się w niej znajdować 2 pliki:
    # 1. Microsoft.AnalysisServices.AdomdClient.dll
    # 2. Microsoft.AnalysisServices.dll
    # Plików szukaj w C:\Windows\assembly\GAC_MSIL\Microsoft.AnalysisServices.(nazwa pliku)\

    :param query: here are required dax queries we need to download data
    :return:
    """
    connStr = "Provider=MSOLAP;Data Source=LB-P-WE-AS;Catalog=PEPCODW"

    conn = Pyadomd(connStr)
    conn.open()
    cursor = conn.cursor()
    cursor.execute(query)
    col_names = [i[0] for i in cursor.description]
    cursor.arraysize = 5000
    df = pd.DataFrame(cursor.fetchall(), columns=col_names)
    conn.close()

    return df

dax_query_list = [daxQ.grading,daxQ.sku_plu,daxQ.md,daxQ.plu_available,
                     daxQ.promo_reg,daxQ.promo_tv,daxQ.prh_data,daxQ.perf_dep,daxQ.pcal,daxQ.prh]

def runDaxQueries(startWeek,endWeek,dep,dax_query_list,path):

    wb = xw.Book.caller()
    ws = wb.sheets["QueryPar2"]

    ws.range("G3:H20").clear_contents() # clear cells with times
    ws.range("K3:L20").clear_contents() # clear cells with info about the files

    for index, query in enumerate(dax_query_list):

        ws.range("C20").value = dep
        ws.range("C21").value = query.__name__
        ws.range("C23").value = "Not Ready"
        ws.range((index + 3, 7)).value = datetime.datetime.now()

        df = runDaxQuery(startWeek,endWeek,dep,query,path)
        q_name = query.__name__ + '.csv'
        df.to_csv(path + q_name.capitalize(), index=False)

        ws.range((index + 3, 8)).value = datetime.datetime.now()
        ws.range((index + 3, 11)).value = query.__name__
        ws.range("C23").value = "Ready"

def runDaxQueriesExe(startWeek,endWeek,dep,dax_query_list,path, hierIdx):
    if type(dep) == list:
        for index, query in enumerate(dax_query_list):
            df_temp = pd.DataFrame()
            for dep_index, dep_name in enumerate(dep):
                startQuery = datetime.datetime.now()
                startQuery = startQuery.strftime("%H:%M:%S")  # startQuery.strftime("%d/%m/%Y %H:%M:%S")
                showStatus(path, 1, dep_name, hierIdx[dep_index], week=None)
                try:
                    df = runDaxQuery(startWeek, endWeek, dep_name, query, path)
                except:
                    try:
                        time.sleep(60)
                        df = runDaxQuery(startWeek, endWeek, dep_name, query, path)
                    except Exception as ex:
                        file = open(path + "ConnectionIssue.txt", "w")
                        file.write(f"Connection error:\n{ex}\n")
                        file.close()
                        return None

                df['HierarchyIndex'] = hierIdx[dep_index]
                df_temp = df_temp.append(df)
                endQuery = datetime.datetime.now()
                endQuery = endQuery.strftime("%H:%M:%S")
                saveLog(startQuery, endQuery, dep_name, hierIdx[dep_index], week=None, path=path, query=query.__name__)

            q_name = query.__name__ + '.csv'
            df_temp.to_csv(path + q_name.capitalize(), index=False)
    else:
        for index, query in enumerate(dax_query_list):
            startQuery = datetime.datetime.now()
            startQuery = startQuery.strftime("%H:%M:%S")  # startQuery.strftime("%d/%m/%Y %H:%M:%S")
            showStatus(path, 1, dep, hierIdx, week=None)

            df_temp = pd.DataFrame()
            try:
                df = runDaxQuery(startWeek, endWeek, dep, query, path)
            except:
                try:
                    time.sleep(60)
                    df = runDaxQuery(startWeek, endWeek, dep, query, path)
                except Exception as ex:
                    file = open(path + "ConnectionIssue.txt", "w")
                    file.write(f"Connection error:\n{ex}\n")
                    file.close()
                    return None

            df['HierarchyIndex'] = hierIdx
            df_temp = df_temp.append(df)
            endQuery = datetime.datetime.now()
            endQuery = endQuery.strftime("%H:%M:%S")
            saveLog(startQuery, endQuery, dep, hierIdx, week=None, path=path, query=query.__name__)

        q_name = query.__name__ + '.csv'
        df_temp.to_csv(path + q_name.capitalize(), index=False)

    showStatus(path, 0, dep, hier=None, week=None) # 0 means Ready

def runDaxQuery(startWeek,endWeek,dep,dax_query,path):
    df = dataFrameFromTabular(dax_query(startWeek, endWeek, dep))

    return df

def weeksCalculation(startWeek, endWeek, maxWeekNo):
    """
    Comments:
    NOT USED
        W raporcie pcal znajdują się wszystkie tygodnie.
        :param startWeek:
        :param endWeek:
        :param maxWeekNo: number of weeks in the year from startWeek variable
        :return:
    """

    list_of_weeks = []
    if int(str(startWeek)[-2:]) > int(str(endWeek)[-2:]):
        weekNumbers = maxWeekNo - int(str(startWeek)[-2:]) + int(str(endWeek)[-2:])
    else:
        weekNumbers = endWeek - startWeek

    weekRatio = 0
    for weeks in range(0,weekNumbers + 1):
        week = startWeek + weeks - weekRatio

        if int(str(week)[-2:]) == 52:
            year = int(str(week)[:4]) + 1
            startWeek = int(str(year) + "01")
            weekRatio = weeks + 1

        list_of_weeks.append(week)
        #print(week)
    return list_of_weeks

def getExcelData():
    wb = xw.Book.caller()

    # Downloader (QueryPar)
    ws = wb.sheets['QueryPar2'] # wb.sheets[3]
    dep = ws["C4"].value
    startWeek = ws["T3"].value
    endWeek = ws["T4"].value
    path = ws["C5"].value + "\\"

    # Calculations (py_inputs)
    ws2 = wb.sheets['py_inputs']
    fileName = ws2["O20"].value
    chosenHierarchy = ws2["O5"].value
    MinPar = ws2["O14"].value # or O24??
    eDur = ws2["O25"].value
    wDur = ws2["O26"].value
    sDur = ws2["O27"].value

    return dep, startWeek,endWeek, path, fileName, chosenHierarchy, MinPar, eDur, wDur, sDur

def getCalcParam(path,fileName):
    wb = xw.Book.caller()
    ws = wb.sheets["QueryPar2"]

    ws.range("C21").value = "parameters"
    ws.range("C23").value = "Not Ready"
    ws.range("G13").value = datetime.datetime.now()

    df = pd.read_excel(fileName, sheet_name='py_inputs', skiprows=1,
                       usecols=['Variables Header', 'Chosen Variables'],converters={'Variables Header':str,'Chosen Variables':str})
    df = df.dropna()
    df = df[16:]
    df.to_csv(path + "parameters.csv", index=False)

    ws.range("H13").value = datetime.datetime.now()
    ws.range("K13").value = "parameters"
    ws.range("C23").value = "Ready"

def loadParameters():
    """
    Load parameters foe Inventory report. Used by Admin once per quarter
    :return:
    """
    dir_ = r'\\10.2.5.140\zasoby\Planowanie\PERSONAL FOLDERS\Mariusz Borycki\HitMidKit_DataBase'
    f_name = '_main files\HitMidKit_parameters.xlsm'
    directory = dir_ + "\\" + f_name
    df = pd.read_excel(directory, sheet_name='py_inputs', skiprows=1)
    df = df[['Variables Header', 'Chosen Variables']]

    df2 = pd.read_excel(directory, sheet_name='QueryPar', skiprows=1)
    df2 = df2[['Chosen Hierarchy', 'Index']]

    startWeek = int(df[df['Variables Header'] == 'Week From']['Chosen Variables'])
    endWeek = int(df[df['Variables Header'] == 'Week To']['Chosen Variables'])
    minPar = int(df[df['Variables Header'] == 'Minimum Sales Units']['Chosen Variables'])
    dep = df[df['Variables Header'] == 'Departament']['Chosen Variables'].item()
    departmentList = list(df2['Chosen Hierarchy'].unique())
    departmentList = [x for x in departmentList if str(x) != 'nan']
    hierIdx = df[df['Variables Header'] == 'Hierarchy']['Chosen Variables'].item()
    indexList = list(df2['Index'].unique())
    indexList = [x for x in indexList if str(x) != 'nan']
    reportType = df[df['Variables Header'] == 'Report Type']['Chosen Variables'].item()

    path = str(df[df['Variables Header'] == 'Path']['Chosen Variables'].item()) + "\\"

    if dep == 0:
        dep = departmentList
        hierIdx = indexList
    else:
        dep = dep
        hierIdx = hierIdx

    return dep, startWeek, endWeek, minPar, path, hierIdx, reportType

def loadInventory(startWeek,endWeek, minPar, dep, path, hierIdx):
    if type(dep) == list:
        for dep_index, dep_name in enumerate(dep):
            weeksList = []
            wList = dataFrameFromTabular(daxQ.pcal(startWeek, endWeek, dep_name))
            for week in wList.iloc[:, 5]:
                weeksList.append(week)
            del wList

            col = ['SKU PLU', 'SKU Colour', 'STR Number', 'Pl Year', 'Pl Week', 'SalesU', 'SalesV', 'SalesM', 'CSOHU', 'CSOHV']
            inventory_df = pd.DataFrame()
            for week in weeksList:
                startQuery = datetime.datetime.now()
                startQuery = startQuery.strftime("%H:%M:%S")  # startQuery.strftime("%d/%m/%Y %H:%M:%S")
                showStatus(path, 1, dep_name, hierIdx[dep_index], week)
                try:
                    df = dataFrameFromTabular(daxQ.inventory(startWeek, endWeek, week, minPar, dep_name))
                except:
                    try:
                        time.sleep(60)
                        df = dataFrameFromTabular(daxQ.inventory(startWeek, endWeek, week, minPar, dep_name))
                    except Exception as ex:
                        file = open(path + "ConnectionIssue.txt", "w")
                        file.write(f"Connection error:\n{ex}\n")
                        file.close()
                        return None

                df['total'] = np.where(df.iloc[:, 5:].sum(axis=1) > 0, 1, 0)
                df = df[df.total > 0]
                df.drop(columns={'total'}, inplace=True)

                inventory_df = inventory_df.append(df)
                endQuery = datetime.datetime.now()
                endQuery = endQuery.strftime("%H:%M:%S")
                saveLog(startQuery, endQuery, dep_name, hierIdx[dep_index], week, path, query='Inventory')

            inventory_df.columns = col
            inventory_df.to_csv(path + f"{hierIdx[dep_index]}_HitMidKit.csv", index=False)
    else:
        weeksList = []
        wList = dataFrameFromTabular(daxQ.pcal(startWeek, endWeek, dep))
        for week in wList.iloc[:, 5]:
            weeksList.append(week)
        del wList

        col = ['SKU PLU','SKU Colour','STR Number','Pl Year','Pl Week','SalesU','SalesV','SalesM','CSOHU','CSOHV']
        inventory_df = pd.DataFrame()
        for week in weeksList:
            startQuery = datetime.datetime.now()
            startQuery = startQuery.strftime("%H:%M:%S")  # startQuery.strftime("%d/%m/%Y %H:%M:%S")
            showStatus(path, 1, dep, hierIdx, week)
            try:
                df = dataFrameFromTabular(daxQ.inventory(startWeek, endWeek, week, minPar, dep))
            except:
                try:
                    time.sleep(60)
                    df = dataFrameFromTabular(daxQ.inventory(startWeek, endWeek, week, minPar, dep))
                except Exception as ex:
                    file = open(path + "ConnectionIssue.txt", "w")
                    file.write(f"Connection error:\n{ex}\n")
                    file.close()
                    return None
            df['total'] = np.where(df.iloc[:, 5:].sum(axis=1) > 0, 1, 0)
            df = df[df.total > 0]
            df.drop(columns={'total'}, inplace=True)
            inventory_df = inventory_df.append(df)

            endQuery = datetime.datetime.now()
            endQuery = endQuery.strftime("%H:%M:%S")
            saveLog(startQuery, endQuery, dep, hierIdx, week, path, query='Inventory')

        inventory_df.columns = col
        inventory_df.to_csv(path + f"{hierIdx}_HitMidKit.csv", index=False)

    showStatus(path, 0, dep=None, hier=None, week=None)  # 0 status means: NOT busy. Inventory file is ready

def showStatus(path,y_n,dep,hier,week):
    if y_n == 1:
        if os.path.exists(path + "Ready.txt"):
            os.remove(path + "Ready.txt")

        file = open(path + "Not Ready.txt", "w")
        file.write(f"Data are still downloading...\n\n"
                   f"Department: {dep}\n"
                   f"Hierarchy: {hier}\n"
                   f"Week: {week}")
        file.close()
    else:
        if os.path.exists(path + "Not Ready.txt"):
            os.remove(path + "Not Ready.txt")

        file = open(path + "Ready.txt", "w")
        file.write("Data have been downloaded")

        file.close()

def saveLog(startQuery, endQuery, dep, hierIdx, week, path, query):
    today = datetime.datetime.today()
    today = today.strftime("%d/%m/%Y")
    queryTime = str(datetime.datetime.strptime(endQuery, "%H:%M:%S") - datetime.datetime.strptime(startQuery, "%H:%M:%S"))
    logList = [today, query, dep, str(hierIdx), str(week), startQuery, endQuery, queryTime]

    with open(path + '../Logs/Log_detailed_HitMidKit.log', 'a+') as log_file:
        log_file.write("\n")
        log_file.writelines(';'.join(logList))


"""
Hit Mid Kit Model Functions
"""
def dataFrameFromTabular(query):
    """
    # w path.append jest istotna ścieżka
    # muszą się w niej znajdować 2 pliki:
    # 1. Microsoft.AnalysisServices.AdomdClient.dll
    # 2. Microsoft.AnalysisServices.dll
    # Plików szukaj w C:\Windows\assembly\GAC_MSIL\Microsoft.AnalysisServices.(nazwa pliku)\

    :param query: here are required dax queries we need to download data
    :return:
    """
    connStr = "Provider=MSOLAP;Data Source=LB-P-WE-AS;Catalog=PEPCODW"

    conn = Pyadomd(connStr)
    conn.open()
    cursor = conn.cursor()
    cursor.execute(query)
    col_names = [i[0] for i in cursor.description]
    cursor.arraysize = 5000
    df = pd.DataFrame(cursor.fetchall(), columns=col_names)
    conn.close()

    return df

def changeColumnName(columns_names):
    sub_a = '['  # substring to be searched
    sub_b = ']'
    nums_a = columns_names.str.find(sub_a)  # creating and passing series to new column
    nums_b = columns_names.str.find(sub_b)
    cols = columns_names

    cols_new = list()
    for idx, col in enumerate(cols):
        num_a = nums_a[idx] + 1
        num_b = nums_b[idx]
        num_b = 20 if num_b < 0 else num_b

        cols_new.append(col[num_a:num_b])
    return cols_new

def createOption(df):
    df['Option'] = df['SKU PLU'].astype(str) + " " + df['SKU Colour'].astype(str)
    return df

def calcGrading(df, g2, g3, ans):
    df = df.sort_values('DepSalesV', ascending=False)
    df['RankX'] = np.arange(df.shape[0]) + 1

    storeCount = len(df['STR Number'].unique())
    df['Cumm%'] = np.where(ans == "Count", df['RankX'] / storeCount,
                           df['DepSalesV'].cumsum() / df['DepSalesV'].sum())

    df['Grade'] = np.where(df['Cumm%'] < float(g3), 3, np.where(df['Cumm%'] < float(g2), 2, 1))
    return df

def addCountry(df_inv, df_grad, grade2, grade3, var_answer):
    # STR Company + Grade
    df = calcGrading(df_grad, grade2, grade3, var_answer)
    df = df[['STR Number', 'Grade', 'STR Company']]
    df['STR Number'] = df['STR Number'].astype(int)

    df2 = pd.merge(df_inv, df, on='STR Number', how='left')
    df2.rename(columns={'Grade': 'StoreGrade'}, inplace=True)

    return df2

def addWeekKey(df_inv, df_pcal):
    # Wk_Key
    df1 = df_pcal[['Pl Year', 'Pl Week', 'Wk_Key']].drop_duplicates()
    df1['Pl Year'] = df1['Pl Year'].astype(int)

    df2 = pd.merge(df_inv, df1, on=['Pl Year', 'Pl Week'], how='inner')
    df2.drop(columns={'Pl Year', 'Pl Week'}, inplace=True)

    cols = ['STR Number', 'STR Company', 'StoreGrade', 'Option', 'Multi',
            'Wk_Key', 'SalesU', 'SalesV', 'CSOHU', 'CSOHV']
    df2 = df2[cols]

    return df2

def showPromo(df_inv, df_promo, df_sku):
    """
    show promo start and promo end just for available PLUs per country
    """

    df_pro = df_promo[['Option', 'STR Company', 'Wk_Key']].drop_duplicates()
    df_pro['PeriodType'] = 'Promo'
    df2 = pd.merge(df_pro, df_inv, on=['Option', 'STR Company', 'Wk_Key'], how='right')

    df3 = df_sku[['Option', 'SKU PPL Initial Retail Price']].drop_duplicates()
    df3 = pd.merge(df3, df2, on=['Option'], how='right')

    df3['stockAUP'] = np.where((df3['CSOHV'] / df3['CSOHU']).isnull(),
                               df3['SalesV'] / df3['SalesU'],
                               (df3['CSOHV'] / df3['CSOHU']))

    df3['PeriodType'] = np.where(df3['PeriodType'] == 'Promo',
                                 df3['PeriodType'],
                                 np.where(df3['stockAUP'] >= df3['SKU PPL Initial Retail Price'] * 0.85,
                                          "Regular",
                                          "Markdown"))

    cols = ['STR Number', 'STR Company', 'StoreGrade', 'Option', 'Multi', 'Available', 'SKU PPL Initial Retail Price',
            'Wk_Key', 'SalesU', 'SalesV', 'CSOHU', 'CSOHV', 'PeriodType']  # ,'Promo Start','Promo End'
    df3 = df3[cols]

    # Promo dates
    df_promo = df3.loc[df3.PeriodType == 'Promo', ('Option', 'Wk_Key')].drop_duplicates()
    df_promo['Promo Start'] = df_promo['Wk_Key']
    df_promo['Promo End'] = df_promo['Wk_Key']

    df_promo.drop(columns={'Wk_Key'}, inplace=True)
    df_promo = df_promo.groupby('Option').agg({'Promo Start': 'min', 'Promo End': 'max'}).reset_index()

    df4 = pd.merge(df3, df_promo, on='Option', how='left')

    df4['Promo Start'].fillna(0, inplace=True)
    df4['Promo Start'] = df4['Promo Start'].astype(int)

    df4['Promo End'].fillna(0, inplace=True)
    df4['Promo End'] = df4['Promo End'].astype(int)

    # Start and End cannot be equal
    # df4['Promo End'] = np.where(df4['Promo End']==0,0,
    # np.where(df4['Promo End']==df4['Promo Start'],df4['Promo End'] + 1, df4['Promo End']))

    df4['Promo End'] = np.where(df4['Promo End'] == 0, 0, df4['Promo End'])

    return df4

def replace_grade(grade):
    if grade not in ['1', '2', '3']:
        return '0'
    else:
        return grade

def GradeStores(Plu_available, Sku_plu, Grade2, Grade3, Grading, var_answer):
    """
    StoresCount: Amount of stores in reality (based od Grading)
    GradeStores: Amount of stores based on 'Cut Off - variable'

    """

    df = Sku_plu[['Option', 'SKU Store Grade']].drop_duplicates()
    df2 = pd.merge(Plu_available, df, on='Option', how='inner')  # czy dać tu inner czy outer??
    df2 = df2[df2['Available'] == 1]  # czy dać tu inner czy outer??
    # df2['SKU Store Grade'] = df2['SKU Store Grade'].astype(str)

    df3 = calcGrading(Grading, Grade2, Grade3, var_answer).groupby(['Grade', 'STR Company'])['STR Number'].count() \
        .reset_index().groupby('STR Company')['STR Number'].sum().reset_index()

    df4 = pd.merge(df2, df3, on='STR Company', how='inner')
    df4.rename(columns={'STR Number': 'StoresCount', 'SKU Store Grade': 'Plu_Grade'}, inplace=True)

    df4['GradeStores'] = np.where(df4.Plu_Grade == 2,
                                  df4.StoresCount * Grade2,
                                  np.where(df4.Plu_Grade == 3,
                                           df4.StoresCount * Grade3,
                                           df4.StoresCount))

    df4 = df4[['Option', 'STR Company', 'Plu_Grade', 'StoresCount', 'GradeStores']]
    # Plu_stores_count = Plu_available.groupby(['Option','Plu_Grade'])['GradeStores'].sum().reset_index()

    df5 = df4.groupby('Option')['GradeStores'].sum().reset_index()

    return df5

def effectiveWeeks(df):
    df_temp = df[(df.PeriodType != 'Markdown') & (df.Wk_Key >= df.WeekFrom) & (df.Wk_Key <= df.WeekTo) & (
                df.StoreGrade >= df['SKU Store Grade'])] \
        .groupby(['Option', 'Wk_Key', 'SKU Store Grade']).agg({'MinInStock': 'sum', 'MinStoresX': 'mean'}).reset_index()
    df_temp['Effective Weeks'] = np.where(df_temp.MinInStock >= df_temp.MinStoresX, 1, 0)
    df_temp = df_temp[df_temp['Effective Weeks'] == 1]
    df_temp = df_temp.groupby('Option')['Effective Weeks'].sum().reset_index()
    return df_temp

def weeksCalc(df, df_sku, WeekMin, WeekMax, eDur, wDur, sDur):
    df1 = df.groupby(['Option', 'SKU Store Grade']).agg({'SalesU': 'sum', 'SalesV': 'sum', 'InStock': 'sum'}).reset_index()
    df1.rename(columns={'InStock': 'In-Stock Stores'}, inplace=True)

    # Sprzedaż dla wszystkich ale InStock per grade?? Czy to jest poprawna formuła w DataModel dla Sales Units in Period??
    df_temp1 = df[(df.PeriodType != 'Markdown')].groupby(['Option', 'Wk_Key', 'PeriodType']) \
        .agg({'MinStores': 'mean'}).reset_index()

    df_temp2 = df[(df.PeriodType != 'Markdown') & (df.StoreGrade >= df['SKU Store Grade'])].groupby(['Option', 'Wk_Key', 'PeriodType']) \
        .agg({'InStock': 'sum'}).reset_index()

    # Total Sales with Exclusion
    df_temp3 = pd.merge(df_temp1, df_temp2, on=['Option', 'Wk_Key', 'PeriodType'], how='inner')

    # weeks from/to
    df_temp4 = df_temp3.groupby(['Option', 'Wk_Key']).agg({'InStock': 'sum', 'MinStores': 'mean'}).reset_index()
    df_WeekStartTemp = df_temp4[df_temp4['InStock'] >= df_temp4['MinStores']].groupby('Option')['Wk_Key'].min().reset_index()
    df_WeekStartTemp.rename(columns={'Wk_Key': 'WeekStartTemp'}, inplace=True)

    # left join because if Merch Type = Y then no matter if InStock >= MinStores. Double check it. CONFIRM!
    df_temp5 = df.merge(df_WeekStartTemp, on='Option', how='left')

    df_Weeks = chooseWeeks(df_temp5, df_sku, WeekMin, WeekMax, eDur, wDur, sDur)
    df_temp3 = pd.merge(df_temp5, df_Weeks, on='Option', how='left')
    df_temp3.drop(columns={'WeekStartTemp'}, inplace=True)

    return df_temp3

def chooseWeeks(df_inv, df_sku, WeekMin, WeekMax, eDur, wDur, sDur):
    df = df_inv[['Option', 'WeekStartTemp']].drop_duplicates()
    df2 = df_sku.merge(df, on='Option', how='inner')

    # choose the weeks for summary the data (entire period 1st step without stores exclusion)
    df2['WeekStart'] = np.where(df2['WeekStartTemp'] == 0,
                                0,
                                np.where(df2['WeekStartTemp'] == 1,
                                         1,
                                         df2['WeekStartTemp'] + 1))

    df2['WeekFrom'] = np.where(np.where(((df2['SKU Merch Group'] == 'Y') & (df2['SKU Merch Season Type'] == 'A')),
                                        WeekMin,
                                        df2['WeekStart']) > WeekMax,
                               0,
                               np.where(((df2['SKU Merch Group'] == 'Y') & (df2['SKU Merch Season Type'] == 'A')),
                                        WeekMin,
                                        df2['WeekStart']))

    df2['WeekTo'] = np.where(df2['SKU Merch Group'] == 'Y',
                             WeekMax,
                             np.where(df2['SKU Merch Group'] == 'E',
                                      df2['WeekFrom'] + eDur - 1,
                                      np.where(df2['SKU Merch Group'] == 'W',
                                               df2['WeekFrom'] + wDur - 1,
                                               df2['WeekFrom'] + sDur - 1)))

    df2.fillna(0, inplace=True)
    df2['WeekFrom'] = df2['WeekFrom'].astype(int)
    df2['WeekTo'] = df2['WeekTo'].astype(int)

    df2['WeekTo'] = np.where(df2['WeekFrom'] == 0,
                             0,
                             np.where(df2['WeekTo'] > WeekMax,
                                      WeekMax,
                                      df2['WeekTo']))

    df2 = df2[['Option', 'WeekFrom', 'WeekTo']].drop_duplicates()

    return df2

def grade_multiEquivalentU(df_inv, df_sku, var_MinTotSls):
    # Dictioneries
    avgROS = dict({'1': [], '2': [], '3': []})
    avgROS_Ratio = dict({'1': [], '2': [], '3': []})

    # Filtering only SKU PLU GRADE = 1 (sells in each store)
    df_sku = df_sku.loc[(df_sku['SKU Store Grade'] == 1), 'Option'].drop_duplicates().to_frame()

    # total sales per Option
    df_sales = df_inv.groupby('Option')['SalesU'].sum().reset_index()

    # Sales >= target (MinTotSls) per Option from df_sku
    df = pd.merge(df_sku, df_sales, on='Option', how='left')
    df.rename(columns={'SalesU': 'SalesTemp'}, inplace=True)
    df['SalesTemp'].fillna(0, inplace=True)
    df = df[df['SalesTemp'] >= var_MinTotSls]  # stores with minimum sales target per SKU

    for grade in range(1, 4):
        # Ros Per Grade calculation: Total sales and Amount of stores calculated for each Grade per Option
        df_grade = df_inv[df_inv.StoreGrade >= grade].groupby('Option').agg(
            {'STR Number': 'count', 'SalesU': 'sum'}).reset_index()
        df_grade['ROS_PerGrade'] = df_grade['SalesU'] / df_grade['STR Number']

        # Grades just for a chosen Options (with targeted sales)
        df_temp = pd.merge(df, df_grade[['Option', 'ROS_PerGrade']], on='Option', how='inner')

        # Appending grades into a dict
        avgROS[str(grade)].append(df_temp['ROS_PerGrade'].mean())

    # Calculate ratio per grades
    for x in range(1, 4):
        result = avgROS[str(x)][0] / avgROS['1'][0]
        avgROS_Ratio[str(x)].append(result)

    return avgROS, avgROS_Ratio

def grade_multiEquivalentV(df_inv, df_sku, var_MinTotSls):
    # Dictioneries
    avgROS = dict({'1': [], '2': [], '3': []})
    avgROS_Ratio = dict({'1': [], '2': [], '3': []})

    # Filtering only SKU PLU GRADE = 1 (sells in each store)
    df_sku = df_sku.loc[(df_sku['SKU Store Grade'] == 1), 'Option'].drop_duplicates().to_frame()

    # total sales per Option
    df_sales = df_inv.groupby('Option').agg({'SalesU': 'sum', 'SalesV': 'sum'}).reset_index()

    # Sales >= target (MinTotSls) per Option from df_sku
    df = pd.merge(df_sku, df_sales, on='Option', how='left')
    df.rename(columns={'SalesV': 'SalesTemp'}, inplace=True)
    df['SalesTemp'].fillna(0, inplace=True)
    df = df[df['SalesU'] >= var_MinTotSls]  # stores with minimum sales target per SKU
    df.drop(columns={'SalesU'}, inplace=True)

    for grade in range(1, 4):
        # Ros Per Grade calculation: Total sales and Amount of stores calculated for each Grade per Option
        df_grade = df_inv[df_inv.StoreGrade >= grade].groupby('Option').agg(
            {'STR Number': 'count', 'SalesV': 'sum'}).reset_index()
        df_grade['ROS_PerGrade'] = df_grade['SalesV'] / df_grade['STR Number']

        # Grades just for a chosen Options (with targeted sales)
        df_temp = pd.merge(df, df_grade[['Option', 'ROS_PerGrade']], on='Option', how='inner')

        # Appending grades into a dict
        avgROS[str(grade)].append(df_temp['ROS_PerGrade'].mean())

    # Calculate ratio per grades
    for x in range(1, 4):
        result = avgROS[str(x)][0] / avgROS['1'][0]
        avgROS_Ratio[str(x)].append(result)

    return avgROS, avgROS_Ratio

def avgSellThru(df, df_sku, WeekMin, WeekMax, eDur, wDur, sDur):
    cols = ['Option', 'Avg Sell-Through In Period', 'SKU Store Grade']
    df_final = pd.DataFrame(columns=cols)

    for grade in range(1, 4):
        df_table = df[(df.StoreGrade >= grade)].copy()  # &(df4.PeriodType!='Markdown')
        df_temp1 = df_table.groupby(['Option', 'Wk_Key', 'PeriodType']).agg(
            {'SalesU': 'sum', 'CSOHU': 'sum', 'InStock': 'sum',
             'MinStores': 'mean', 'MinStoresX': 'mean'}).reset_index()

        # weeks from/to
        df_temp2 = df_table.groupby(['Option', 'Wk_Key']).agg({'InStock': 'sum', 'MinStores': 'mean'}).reset_index()
        df_WeekStartTemp = df_temp2[df_temp2['InStock'] >= df_temp2['MinStores']].groupby('Option')[
            'Wk_Key'].min().reset_index()
        df_WeekStartTemp.rename(columns={'Wk_Key': 'WeekStartTemp'}, inplace=True)
        df_temp3 = pd.merge(df_temp1, df_WeekStartTemp, on='Option', how='left')

        df_weeks = chooseWeeks(df_temp3, df_sku, WeekMin, WeekMax, eDur, wDur, sDur)
        df_temp4 = pd.merge(df_temp3, df_weeks, on='Option', how='left')
        df_temp5 = df_temp4[((df_temp4['Wk_Key'] >= df_temp4['WeekFrom']) &
                             (df_temp4['Wk_Key'] <= df_temp4['WeekTo']))]. \
            groupby(['Option', 'Wk_Key']).agg(
            {'SalesU': 'sum', 'CSOHU': 'sum', 'InStock': 'sum', 'MinStoresX': 'mean'}).reset_index()

        df_temp5['Avg Sell-Through In Period'] = np.where(df_temp5.InStock >= df_temp5.MinStoresX,
                                                          df_temp5.SalesU / (df_temp5.CSOHU + df_temp5.SalesU),
                                                          0)

        df_temp5 = df_temp5[df_temp5['Avg Sell-Through In Period'] > 0]
        df_temp5 = df_temp5.groupby('Option')['Avg Sell-Through In Period'].mean().reset_index()
        df_temp5['SKU Store Grade'] = grade
        df_final = df_final.append(df_temp5, ignore_index=True)

    return df_final

def ROS_ProdGrade(df_inv, df_sku, WeekMin, WeekMax, eDur, wDur, sDur, grade):
    # total
    df_table = df_inv[df_inv.StoreGrade >= grade].copy()

    df_final = df_table.groupby(['Option', 'SKU Store Grade']).agg(
        {'SalesU': 'sum', 'SalesV': 'sum', 'InStock': 'sum'}).reset_index()
    df_final.rename(columns={'SalesU': 'Sales Units', 'SalesV': 'Sales Value', 'InStock': 'In-Stock Stores'},
                    inplace=True)

    # Sprzedaż dla wszystkich ale InStock per grade?? Czy to jest poprawna formuła w DataModel dla Sales Units in Period??
    df_temp1 = df_table[(df_table.PeriodType != 'Markdown')].groupby(['Option', 'Wk_Key', 'PeriodType']) \
        .agg({'MinStoresX': 'mean', 'MinStores': 'mean', 'SalesU': 'sum', 'SalesV': 'sum'}).reset_index()

    df_temp2 = df_table[
        (df_table.PeriodType != 'Markdown') & (df_table.StoreGrade >= df_table['SKU Store Grade'])].groupby(
        ['Option', 'Wk_Key', 'PeriodType']) \
        .agg({'InStock': 'sum'}).reset_index()

    df_temp3 = df_table[(df_table.PeriodType != 'Markdown')].groupby(['Option', 'Wk_Key', 'PeriodType']) \
        .agg({'MinInStock': 'sum'}).reset_index()

    # Total Sales with Exclusion
    df2 = pd.merge(df_temp1, df_temp2, on=['Option', 'Wk_Key', 'PeriodType'], how='inner')
    df2 = pd.merge(df2, df_temp3, on=['Option', 'Wk_Key', 'PeriodType'], how='inner')

    df2['SalesUX'] = np.where(df2['InStock'] >= df2['MinStoresX'], df2['SalesU'], 0)
    df2['SalesVX'] = np.where(df2['InStock'] >= df2['MinStoresX'], df2['SalesV'], 0)
    df2['MinInStock_X'] = np.where(df2['InStock'] >= df2['MinStoresX'], df2['MinInStock'], 0)  # TempInvent (PowerPivot)

    # weeks from/to
    df_temp4 = df2.groupby(['Option', 'Wk_Key']).agg({'InStock': 'sum', 'MinStores': 'mean'}).reset_index()
    df_WeekStartTemp = df_temp4[df_temp4['InStock'] >= df_temp4['MinStores']].groupby('Option')[
        'Wk_Key'].min().reset_index()
    df_WeekStartTemp.rename(columns={'Wk_Key': 'WeekStartTemp'}, inplace=True)

    # left join because if Merch Type = Y then no matter if InStock >= MinStores. Double check it. CONFIRM!
    df_temp5 = df2.merge(df_WeekStartTemp, on='Option', how='left')

    df_Weeks = chooseWeeks(df_temp5, df_sku, WeekMin, WeekMax, eDur, wDur, sDur)
    df_temp5 = pd.merge(df_temp5, df_Weeks, on='Option', how='left')

    # grouping another
    sourceList = ['MinInStock_X', 'SalesUX', 'SalesVX']

    newNameList = ['In-Stock Stores in Period', 'Sales Units in Period', 'Sales Value in Period']

    for x in zip(sourceList, newNameList):
        df_temp = df_temp5[((df_temp5['Wk_Key'] >= df_temp5['WeekFrom']) &
                            (df_temp5['Wk_Key'] <= df_temp5['WeekTo']))]. \
            groupby(['Option'])[x[0]].sum().reset_index()

        df_temp.rename(columns={x[0]: x[1]}, inplace=True)
        df_final = pd.merge(df_final, df_temp, on='Option', how='left')

    ################################
    # ROS -> sales divided by amount of stores with excluded weeks
    df_final['ROS_ProdGrade'] = df_final['Sales Units in Period'] / df_final[
        'In-Stock Stores in Period']  # MinInStock_X
    df_final['ROS_ProdGradeV'] = df_final['Sales Value in Period'] / df_final[
        'In-Stock Stores in Period']  # MinInStock_X

    df_final = df_final.loc[df_final['ROS_ProdGrade'] > 0, ('Option', 'ROS_ProdGrade', 'ROS_ProdGradeV')]
    df_final['SKU Store Grade'] = grade

    return df_final

def InventoryAggregation(df_inv, df_sku, WeekMin, WeekMax, eDur, wDur, sDur):
    # total
    df1 = df_inv.groupby(['Option', 'SKU Store Grade']).agg(
        {'SalesU': 'sum', 'SalesV': 'sum', 'InStock': 'sum'}).reset_index()
    df1.rename(columns={'SalesU': 'Sales Units', 'SalesV': 'Sales Value', 'InStock': 'In-Stock Stores'}, inplace=True)

    ########################################################################
    # Sprzedaż dla wszystkich ale InStock per grade?? Czy to jest poprawna formuła w DataModel dla Sales Units in Period??
    df_temp1 = df_inv[(df_inv.PeriodType != 'Markdown')].groupby(['Option', 'Wk_Key', 'PeriodType']) \
        .agg({'WeekFrom': 'mean', 'WeekTo': 'mean', 'MinStoresX': 'mean', 'MinInStock': 'sum',
              'SalesU': 'sum', 'SalesV': 'sum', 'CSOHU': 'sum'}).reset_index()

    df_temp2 = df_inv[(df_inv.PeriodType != 'Markdown') & (df_inv.StoreGrade >= df_inv['SKU Store Grade'])] \
        .groupby(['Option', 'Wk_Key', 'PeriodType']).agg({'InStock': 'sum'}).reset_index()

    ########################################################################
    # Total Sales with Exclusion
    df2 = pd.merge(df_temp1, df_temp2, on=['Option', 'Wk_Key', 'PeriodType'], how='inner')
    df2['SalesUX'] = np.where(df2['InStock'] >= df2['MinStoresX'], df2['SalesU'], 0)
    df2['SalesVX'] = np.where(df2['InStock'] >= df2['MinStoresX'], df2['SalesV'], 0)
    # df2['CSOHUX'] = np.where(df2['InStock'] >= df2['MinStoresX'], df2['CSOHU'], 0)
    df2['MinInStock_X'] = np.where(df2['InStock'] >= df2['MinStoresX'], df2['MinInStock'], 0)  # TempInvent (PowerPivot)

    ########################################################################
    # Promo Sales
    df2['SalesUPX'] = np.where(((df2['InStock'] >= df2['MinStoresX']) &
                                (df2['PeriodType'] == 'Promo')), df2['SalesU'], 0)
    df2['SalesVPX'] = np.where(((df2['InStock'] >= df2['MinStoresX']) &
                                (df2['PeriodType'] == 'Promo')), df2['SalesV'], 0)
    df2['MinInStock_PX'] = np.where(((df2['InStock'] >= df2['MinStoresX']) &
                                     (df2['PeriodType'] == 'Promo')), df2['MinInStock'], 0)
    ########################################################################
    # Regular Sales
    df2['SalesURX'] = np.where(((df2['InStock'] >= df2['MinStoresX']) &
                                (df2['PeriodType'] == 'Regular')), df2['SalesU'], 0)
    df2['SalesVRX'] = np.where(((df2['InStock'] >= df2['MinStoresX']) &
                                (df2['PeriodType'] == 'Regular')), df2['SalesV'], 0)
    df2['MinInStock_RX'] = np.where(((df2['InStock'] >= df2['MinStoresX']) &
                                     (df2['PeriodType'] == 'Regular')), df2['MinInStock'], 0)

    ########################################################################
    # Summarize values for a chosen periods
    sourceList = ['MinInStock_X', 'MinInStock_PX', 'MinInStock_RX',
                  'SalesUX', 'SalesUPX', 'SalesURX',
                  'SalesVX', 'SalesVPX', 'SalesVRX']

    newNameList = ['In-Stock Stores in Period', 'In-Stock Stores in Period Promo', 'In-Stock Stores in Period Regular',
                   'Sales Units in Period', 'Sales Units in Period Promo', 'Sales Units in Period Regular',
                   'Sales Value in Period', 'Sales Value in Period Promo', 'Sales Value in Period Regular']

    for x in zip(sourceList, newNameList):
        df_temp = df2[((df2['Wk_Key'] >= df2['WeekFrom']) &
                       (df2['Wk_Key'] <= df2['WeekTo']))].groupby(['Option'])[x[0]].sum().reset_index()

        df_temp.rename(columns={x[0]: x[1]}, inplace=True)
        df1 = pd.merge(df1, df_temp, on='Option', how='left')

    ########################################################################
    # Stock
    df_temp = df_inv.loc[df_inv['Wk_Key'] == WeekMax, ('Option', 'CSOHU')].groupby(['Option'])[
        'CSOHU'].sum().reset_index()
    df_temp.rename(columns={'CSOHU': 'Stock Units in Period'}, inplace=True)
    df1 = pd.merge(df1, df_temp, on='Option', how='left')

    ########################################################################
    # Adding average numbers from default table
    df_temp = df_inv.groupby('Option').agg(
        {'Multi': 'mean', 'GradeStores': 'mean', 'MinStores': 'mean', 'MinStoresX': 'mean',
         'WeekFrom': 'mean', 'WeekTo': 'mean', 'Promo Start': 'mean', 'Promo End': 'mean'}).reset_index()
    df1 = pd.merge(df1, df_temp, on='Option', how='left')
    df1.fillna(0, inplace=True)

    ########################################################################
    # Change datatypes
    df1.rename(columns={'WeekFrom': 'Week Start', 'WeekTo': 'Week End'}, inplace=True)
    df1['Week Start'] = df1['Week Start'].astype(int)
    df1['Week End'] = df1['Week End'].astype(int)
    df1['Promo Start'] = df1['Promo Start'].astype(int)
    df1['Promo End'] = df1['Promo End'].astype(int)
    df1['Multi'] = df1['Multi'].astype(int)

    ########################################################################
    #
    df1['% Total Sales Value'] = df1['Sales Value in Period'] / df1['Sales Value']
    df1['% Total Sales Units'] = df1['Sales Units in Period'] / df1['Sales Units']

    ########################################################################
    # ROS -> sales divided by amount of stores with excluded weeks
    df1['RAW ROS_U'] = df1['Sales Units in Period'] / df1['In-Stock Stores in Period']  # MinInStock_X
    df1['RAW ROS_V'] = df1['Sales Value in Period'] / df1['In-Stock Stores in Period']  # MinInStock_X

    df1.fillna(0, inplace=True)

    ########################################################################
    # ROS_ProdGrade
    df_ros2 = ROS_ProdGrade(df_inv, df_sku, WeekMin, WeekMax, eDur, wDur, sDur, 2)
    df_ros3 = ROS_ProdGrade(df_inv, df_sku, WeekMin, WeekMax, eDur, wDur, sDur, 3)
    df_ros = pd.concat([df_ros2, df_ros3])

    ########################################################################
    #
    df1 = pd.merge(df1, df_ros, on=['Option', 'SKU Store Grade'], how='left')
    df1['ROS_ProdGrade'] = np.where(df1['SKU Store Grade'] == 1, df1['RAW ROS_U'], df1['ROS_ProdGrade'])
    df1['ROS_ProdGradeV'] = np.where(df1['SKU Store Grade'] == 1, df1['RAW ROS_V'], df1['ROS_ProdGradeV'])
    df1.drop(columns={'RAW ROS_U', 'RAW ROS_V'}, inplace=True)

    ########################################################################
    # Promo / Non Promo ROS
    df1['NonPromo ROS'] = df1['Sales Units in Period Regular'] / df1['In-Stock Stores in Period Regular']
    df1['Promo ROS'] = df1['Sales Units in Period Promo'] / df1['In-Stock Stores in Period Promo']

    ########################################################################
    # PromoFactor
    ROS_PerGrade_Promo = df1[df1['SKU Store Grade'] == 1]['Sales Units in Period Promo'].sum() / \
                         df1[df1['SKU Store Grade'] == 1]['In-Stock Stores in Period Promo'].sum()
    ROS_PerGrade_Regular = df1[df1['SKU Store Grade'] == 1]['Sales Units in Period Regular'].sum() / \
                           df1[df1['SKU Store Grade'] == 1]['In-Stock Stores in Period Regular'].sum()
    PromoFactor = ROS_PerGrade_Promo / ROS_PerGrade_Regular

    ########################################################################
    # Promo Adjusted ROS
    df1['Promo Adjusted ROS'] = np.where(df1['Sales Units in Period Promo'] == 0,
                                         df1['ROS_ProdGrade'],
                                         (df1['Sales Units in Period Promo'] / PromoFactor + df1[
                                             'Sales Units in Period Regular']) / df1['In-Stock Stores in Period'])
    df1['Promo Adjusted ROS V'] = np.where(df1['Sales Value in Period Promo'] == 0,
                                           df1['ROS_ProdGradeV'],
                                           (df1['Sales Value in Period Promo'] / PromoFactor + df1[
                                               'Sales Value in Period Regular']) / df1['In-Stock Stores in Period'])

    return df1

def calcPerf(df, hier):
    df_perf = df.copy()
    df_perf.rename(columns={'PRH Sub Department': 'SKU Sub Department'}, inplace=True)
    df_perf['Sales Perf Min'] = df_perf['Inflows ACT'] / df_perf['Intake AP']

    ################################################
    # Hardline
    if hier[2] == '2':
        df_perf['Sales Perf'] = df_perf['Sales Retail ACT'] / (
                    df_perf['Sales Retail AP'] * df_perf['Sales Perf Min']) - 1

        df_perf['Tier1'] = np.where(df_perf['Sales Perf'] > 0.05, 0.3,
                                    np.where(df_perf['Sales Perf'] > -0.05, 0.25, 0.2))

        df_perf['Tier2'] = np.where(df_perf['Sales Perf'] > 0.05, 0.6,
                                    np.where(df_perf['Sales Perf'] > -0.05, 0.5, 0.4))

        df_perf['Tier3'] = np.where(df_perf['Sales Perf'] > 0.05, 0.8,
                                    np.where(df_perf['Sales Perf'] > -0.05, 0.75, 0.7))

        Sales_Perf_Min = df_perf['Inflows ACT'].sum() / df_perf['Intake AP'].sum()
        Sales_dep_Perf = df_perf['Sales Retail ACT'].sum() / (df_perf['Sales Retail AP'].sum() * Sales_Perf_Min) - 1
    ################################################
    # Clothing
    else:
        df_perf['Sales Perf'] = df_perf['Sales Retail ACT'] / df_perf['Sales Retail AP'] - 1
        df_perf['SlsTier1'] = np.where(df_perf['Sales Perf'] > 0.05, 0.4,
                                       np.where(df_perf['Sales Perf'] > -0.05, 0.3, 0.2))

        df_perf['SlsTier2'] = np.where(df_perf['Sales Perf'] > 0.05, 0.8,
                                       np.where(df_perf['Sales Perf'] > -0.05, 0.65, 0.5))

        Sales_dep_Perf = df_perf['Sales Retail ACT'].sum() / df_perf['Sales Retail AP'].sum() - 1

    ################################################
    # Total ratios (Sell Through & Markdowns)
    MKD_Perf = -df_perf['Markdown ACT'].sum() / df_perf['Markdown AP'].sum() + 1

    MD_Tier1 = np.where(MKD_Perf > 0.05, 0.8, np.where(MKD_Perf >= -0.05, 0.6, 0.4))
    MD_Tier2 = np.where(MKD_Perf > 0.05, 1.5, np.where(MKD_Perf >= -0.05, 1.2, 0.9))

    ST_Tier_1 = np.where(Sales_dep_Perf > 0.05, 0.4, np.where(Sales_dep_Perf >= -0.05, 0.3, 0.2))
    ST_Tier_2 = np.where(Sales_dep_Perf > 0.05, 0.8, np.where(Sales_dep_Perf >= -0.05, 0.65, 0.5))

    return df_perf, ST_Tier_1, ST_Tier_2, MD_Tier1, MD_Tier2

def calcScoringSellThru(df_score, ST_Tier_1, ST_Tier_2):
    colnames = ['Option', 'SKU Merch Type', 'MerchGroup', 'Avg Sell-Through In Period', 'SKU Sub Department',
                'ItemCountMer']
    df1 = pd.DataFrame(columns=colnames)

    for merch in list(df_score['MerchGroup'].unique()):
        df_temp = df_score[df_score.ItemExcl == 0].loc[df_score['MerchGroup'] == merch, colnames] \
            .sort_values(by='Avg Sell-Through In Period', ascending=False).copy()
        df_temp['RankX'] = np.arange(df_temp.shape[0]) + 1

        # df_temp['Cumm%'] = round(df_temp['RankX']/df_temp['RankX'].max(),2)

        df_temp['ST Value'] = np.where(df_temp['RankX'] <= np.ceil(ST_Tier_1 * df_temp['ItemCountMer']), 'Tier1_V',
                                       np.where(df_temp['RankX'] <= np.ceil(ST_Tier_2 * df_temp['ItemCountMer']),
                                                'Tier2_V', '-'))
        df_temp = df_temp[df_temp['ST Value'] != '-']
        df1 = df1.append(df_temp)

    df2 = df1.groupby(['MerchGroup', 'ST Value'])['Avg Sell-Through In Period'].min().reset_index()
    df2 = pd.pivot_table(df2,
                         index='MerchGroup',
                         columns='ST Value',
                         values='Avg Sell-Through In Period',
                         aggfunc='mean').reset_index()

    df = pd.merge(df_score[['Option', 'MerchGroup', 'Avg Sell-Through In Period']], df2, on='MerchGroup', how='inner')
    df['ST Score'] = np.where(df['Avg Sell-Through In Period'] >= df['Tier1_V'], 1,
                              np.where(df['Avg Sell-Through In Period'] >= df['Tier2_V'], 0.5, 0))
    df = df[['Option', 'ST Score']]

    return df

def calcScoringROS(score_variable, score_valueA, score_valueB, df_score, hier):
    """
    score_variable = 'ROS Margin Value'
    score_valueA = 'ST Value'
    score_valueB = 'ST Score'
    """

    ################################################
    # Hardline
    if hier[2] == '2':

        colnames = ['Option', score_variable, 'SKU Sub Department', 'Tier1', 'Tier2', 'Tier3', 'ItemCountDep']
        df1 = pd.DataFrame(columns=colnames)

        for dep in list(df_score['SKU Sub Department'].unique()):
            df_temp = df_score[df_score.ItemExcl == 0].loc[df_score['SKU Sub Department'] == dep, colnames].sort_values(
                score_variable, ascending=False).copy()
            df_temp['RankX'] = np.arange(df_temp.shape[0]) + 1
            # df_temp['Cumm%'] = round(df_temp['RankX']/df_temp['RankX'].max(),2)
            df_temp[score_valueA] = np.where(df_temp['RankX'] <= np.ceil(df_temp['Tier1'] * df_temp['ItemCountDep']),
                                             'Tier1_V',
                                             np.where(df_temp['RankX'] <= np.ceil(
                                                 df_temp['Tier2'] * df_temp['ItemCountDep']), 'Tier2_V',
                                                      np.where(df_temp['RankX'] <= np.ceil(
                                                          df_temp['Tier3'] * df_temp['ItemCountDep']), 'Tier3_V', '-')))

            df_temp = df_temp[df_temp[score_valueA] != '-']
            df1 = df1.append(df_temp)

        df2 = df1.groupby(['SKU Sub Department', score_valueA])[score_variable].min().reset_index()
        df2 = df2[df2[score_valueA] != 0]

        df2 = pd.pivot_table(df2,
                             index='SKU Sub Department',
                             columns=score_valueA,
                             values=score_variable,
                             aggfunc='mean').reset_index()

        df = pd.merge(df_score[['Option', 'SKU Sub Department', score_variable]], df2, on='SKU Sub Department',
                      how='inner')
        df[score_valueB] = np.where(df[score_variable] >= df['Tier1_V'], 1.5,
                                    np.where(df[score_variable] >= df['Tier2_V'], 1,
                                             np.where(df[score_variable] >= df['Tier3_V'], 0.5, 0)))
        df = df[['Option', score_valueB]]

        ################################################
        # Clothing
    else:
        colnames = ['Option', score_variable, 'SKU Sub Department', 'SlsTier1', 'SlsTier2', 'ItemCountDep']
        df1 = pd.DataFrame(columns=colnames)

        for dep in list(df_score['SKU Sub Department'].unique()):
            df_temp = df_score[df_score.ItemExcl == 0].loc[df_score['SKU Sub Department'] == dep, colnames].sort_values(
                score_variable, ascending=False).copy()
            df_temp['RankX'] = np.arange(df_temp.shape[0]) + 1
            # df_temp['Cumm%'] = round(df_temp['RankX']/df_temp['RankX'].max(),2)
            df_temp[score_valueA] = np.where(df_temp['RankX'] <= np.ceil(df_temp['SlsTier1'] * df_temp['ItemCountDep']),
                                             'Tier1_V',
                                             np.where(df_temp['RankX'] <= np.ceil(
                                                 df_temp['SlsTier2'] * df_temp['ItemCountDep']), 'Tier2_V', '-'))

            df_temp = df_temp[df_temp[score_valueA] != '-']
            df1 = df1.append(df_temp)

        df2 = df1.groupby(['SKU Sub Department', score_valueA])[score_variable].min().reset_index()
        df2 = df2[df2[score_valueA] != 0]

        df2 = pd.pivot_table(df2,
                             index='SKU Sub Department',
                             columns=score_valueA,
                             values=score_variable,
                             aggfunc='mean').reset_index()

        df = pd.merge(df_score[['Option', 'SKU Sub Department', score_variable]], df2, on='SKU Sub Department',
                      how='inner')
        df[score_valueB] = np.where(df[score_variable] >= df['Tier1_V'], 1,
                                    np.where(df[score_variable] >= df['Tier2_V'], 0.5, 0))
        df = df[['Option', score_valueB]]

    return df

def calcScoringMD(df_score, MD_Tier1, MD_Tier2):
    """
    score_variable = 'ROS Margin Value'
    score_valueA = 'ST Value'
    score_valueB = 'ST Score'
    """

    colnames = ['Option', 'SKU Merch Type', 'MerchGroup', 'Sales Value', 'Markdown Value', 'SKU Sub Department',
                'ItemCountMer']
    df1 = pd.DataFrame(columns=colnames)

    for merch in list(df_score['MerchGroup'].unique()):
        df_temp = df_score[df_score.ItemExcl == 0].loc[df_score['MerchGroup'] == merch, colnames] \
            .sort_values(by='MerchGroup', ascending=False).copy()
        df_temp['MD_SLS'] = np.where(df_temp['Markdown Value'] > 0, np.where(df_temp['Sales Value'] > 0,
                                                                             df_temp['Markdown Value'] / df_temp[
                                                                                 'Sales Value'], 1), 0)

        df1 = df1.append(df_temp)

    Tier1_V = (df1['Markdown Value'].sum() / df1['Sales Value'].sum()) * MD_Tier1
    Tier2_V = (df1['Markdown Value'].sum() / df1['Sales Value'].sum()) * MD_Tier2

    df1['MD Score'] = np.where(df1['MD_SLS'] <= Tier1_V, 1, np.where(df1['MD_SLS'] <= Tier2_V, 0.5, 0))
    df1 = df1[['Option', 'MD Score', 'MD_SLS']]

    return df1

#######################################################
def getPath():
    if getattr(sys, 'frozen', False):
        SETTINGS_PATH = os.path.dirname(sys.executable)
    else:
        SETTINGS_PATH = os.path.dirname(os.path.abspath(__file__))

    with open(SETTINGS_PATH + "\Settings.txt", "r") as file:
        first_line = file.readline()

    while first_line[-4:] != 'xlsm':
        first_line = first_line[:-1]
    return first_line

def getParameters(MODEL_PATH): #MODEL_PATH

    df = pd.read_excel(MODEL_PATH, sheet_name='py_inputs', skiprows=1)
    df = df[['Variables Header', 'Chosen Variables']]
    param_names = ['Hierarchy','Departament','Grading Type','Week From','Week To','Minimum Sales Units','MinTotSls','MinStrStk','StrCount','WeekExcl',
                 'E Merch Group Duration','W Merch Group Duration','S Merch Group Duration','Path','Path Inventory','Grade2','Grade3']

    param_values = list()
    for idx in range(len(param_names)):
        param_value = df[df['Variables Header'] == param_names[idx]]['Chosen Variables'].item()
        param_values.append(param_value)

    variables = dict(zip(param_names, param_values))

    return variables

def saveSummary(MODEL_PATH, SHEET_NAME, DF):
    wb = xw.Book(MODEL_PATH) # xw.Book.caller()
    ws = wb.sheets[SHEET_NAME]
    ws.range("B2").value = DF

def saveStatus(MODEL_PATH, INSIGHT, VALUE, VALUE_TOTAL):
    STATUS_NO = VALUE + 1

    wb = xw.Book(MODEL_PATH)
    ws = wb.sheets['QueryPar']
    ws.range("E10").value = INSIGHT
    ws.range("E14").value = STATUS_NO/VALUE_TOTAL

    if STATUS_NO == VALUE_TOTAL:
        ws.range("E16").value = 'Ready'
    else:
        ws.range("E16").value = 'Not Ready'

def saveStatusNew(MODEL_PATH, df):
    wb = load_workbook(MODEL_PATH)
    ws = wb["Summary"]
    ws["A1"] = df

    wb.save(MODEL_PATH)


def showStatusModel(path,y_n,dep,hier,status, value,value_total):
    """
    0 status means: NOT busy. Inventory file is ready;
    showStatus(path, 0, dep=None, hier=None, week=None)
    :param path:
    :param y_n:
    :param dep:
    :param hier:
    :param week:
    :return:
    """
    PROGRESS = "{0:.0f}%".format(value/value_total * 100)

    if y_n == 1:
        if os.path.exists(path + "Ready.txt"):
            os.remove(path + "Ready.txt")

        file = open(path + "Not Ready.txt", "w")
        file.write(f"Data are still downloading...\n\n"
                   f"Department: {dep}\n"
                   f"Hierarchy: {hier}\n"
                   f"Status: {status}\n"
                   f"Progress: {PROGRESS}%")
        file.close()
    else:
        if os.path.exists(path + "Not Ready.txt"):
            os.remove(path + "Not Ready.txt")

        file = open(path + "Ready.txt", "w")
        file.write("Data have been downloaded")

        file.close()
##########################
"""
##########################################
OLD Functions#############################

def NotNeeded1():
    wb = xw.Book.caller()
    #ws = wb.sheets["notes"]
    ws = wb.sheets['py_inputs']

    # ws.range("C9").value = "Parameters"
    # ws.range("D9").value = "Not Ready"
    # ws.range("E9").value = datetime.datetime.now()
    fileName = ws["O20"].value

    df = pd.read_excel(fileName, sheet_name='py_inputs', skiprows=1,
                       usecols=['Variables Header', 'Chosen Variables'],converters={'Variables Header':str,'Chosen Variables':str})
    df.dropna(inplace=True)
    df.reset_index(drop=True, inplace=True)

    #ws.range("C11").value = df
    #ws.range("G9").value = df.loc[df['Variables Header']=='Week To','Chosen Variables'].values
    # ws.range("F9").value = datetime.datetime.now()
    # ws.range("D9").value = "Ready"
    return df

##########################################
def showGrades_old(df):
    columnNames = ['Grade', 'Cut-Off (Count / Value)']
    df = df[-3:].copy()
    df.columns = columnNames
    df['Grade'] = df.loc[:, ('Grade')].str.replace('Grade', '').astype(int)

    grades = list()
    for x in df.iloc[:, 1]:
        grades.append(x)

    return float(grades[1]), float(grades[2])
    
"""